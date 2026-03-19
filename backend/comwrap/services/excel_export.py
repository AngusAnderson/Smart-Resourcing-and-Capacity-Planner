from collections import defaultdict, OrderedDict
from dataclasses import dataclass
from datetime import date
from decimal import Decimal

from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

from django.http import HttpResponse
# from django.db.models import Prefetch
# from django.utils.timezone import make_naive
from openpyxl.cell.cell import MergedCell

from comwrap.models import ForecastAllocation  # adjust import

# pivot logic to build export data structure

def month_range_keys(start_month: date, end_month: date):
    """Inclusive months. start_month/end_month should be first-of-month dates."""
    y, m = start_month.year, start_month.month
    end_y, end_m = end_month.year, end_month.month
    out = []
    while (y, m) <= (end_y, end_m):
        out.append((y, m))
        m += 1
        if m == 13:
            m = 1
            y += 1
    return out

def add_months(d: date, n: int) -> date:
    y = d.year + (d.month - 1 + n) // 12
    m = (d.month - 1 + n) % 12 + 1
    return date(y, m, 1)

@dataclass(frozen=True)
class ExportRowKey:
    employee_id: int
    jobcode_id: int

def build_forecast_export(start_month: date, end_month: date):
    """
    Returns:
      months: list[(y,m)]
      employees: ordered list of employee objects encountered
      data: dict[(employee_id, jobcode_id)] -> dict[(y,m)] -> Decimal
      meta: dict[(employee_id, jobcode_id)] -> info (employee, jobcode, description, etc.)
    """
    months = month_range_keys(start_month, end_month)
    end_exclusive = add_months(end_month, 1)

    qs = (
        ForecastAllocation.objects
        .filter(
            forecast__date__gte=start_month,
            forecast__date__lt=end_exclusive,
        )
        .select_related(
            "employee",
            "forecast",
            "forecast__jobCode",
            "forecast__jobCode__replyEntity",
        )
        .order_by("employee__id", "forecast__jobCode__id", "forecast__date")
    )

    data = defaultdict(lambda: defaultdict(Decimal))
    meta = {}
    employee_order = OrderedDict()  # preserves first-seen order

    for fa in qs:
        emp = fa.employee
        fc = fa.forecast
        jc = fc.jobCode
        key = ExportRowKey(emp.id, jc.id)
        month_key = (fc.date.year, fc.date.month)

        employee_order.setdefault(emp.id, emp)

        # sum in case there are multiple rows per same key+month
        data[key][month_key] += (fa.daysAllocated or Decimal("0"))

        if key not in meta:
            meta[key] = {
                "employee": emp,
                "jobcode": jc,
                "forecast": fc,  # optional
                # add any fields you want to print in static columns
            }

    return months, list(employee_order.values()), data, meta


# Excel export logic to write the above data structure into the template and return as response

MONTH_START_COL = 11  # column K

def format_employee_name(emp):
    #Returns: LASTNAME, FirstName

    # Support both snake_case and camelCase field names
    last_raw = getattr(emp, "name".split()[0], None)
    if last_raw is None:
        last_raw = getattr(emp, "name", "")
    first_raw = getattr(emp, "name".split()[0], None)
    if first_raw is None:
        first_raw = getattr(emp, "name", "")

    last = (last_raw or "").upper()
    first = (first_raw or "").capitalize()
    return f"{last}, {first}".strip(", ")

def month_label(y, m):
    return date(y, m, 1).strftime("%b-%y")

def set_cell_value_safe(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        # Find the merged range that contains this cell and write to its top-left
        for r in ws.merged_cells.ranges:
            if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col:
                ws.cell(row=r.min_row, column=r.min_col).value = value
                return
        # fallback (shouldn't happen)
        return
    cell.value = value


def excel_safe(value):
    """Convert Django/model objects into Excel-safe primitives."""
    if value is None:
        return ""
    # crude check: Django model instances have _meta
    if hasattr(value, "_meta"):
        return str(value)
    return value


def snapshot_row_style(ws, src_row: int, max_col: int):
    """Capture copied styles for a whole row so we can apply them after insert_rows()."""
    styles = []
    for c in range(1, max_col + 1):
        cell = ws.cell(src_row, c)
        styles.append({
            "_style": copy(cell._style),
            "number_format": cell.number_format,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "protection": copy(cell.protection),
        })
    row_height = ws.row_dimensions[src_row].height
    return styles, row_height


def apply_row_style(ws, dst_row: int, styles, row_height=None):
    for c, s in enumerate(styles, start=1):
        cell = ws.cell(dst_row, c)
        cell._style = copy(s["_style"])
        cell.number_format = s["number_format"]
        cell.font = copy(s["font"])
        cell.fill = copy(s["fill"])
        cell.border = copy(s["border"])
        cell.alignment = copy(s["alignment"])
        cell.protection = copy(s["protection"])
    if row_height is not None:
        ws.row_dimensions[dst_row].height = row_height

def export_forecast_xlsx(start_month: date, end_month: date):
    months, employees, data, meta = build_forecast_export(start_month, end_month)

    BASE_DIR = Path(__file__).resolve().parent.parent
    TEMPLATE_PATH = BASE_DIR / "resources" / "excel-templates" / "forecast-template.xlsx"
    wb = load_workbook(TEMPLATE_PATH)
    #select sheet depending on if time range is multi-month or single month
    if len(months) > 1:
        ws = wb["Multi-Month Template"]
    else:
        ws = wb["Single-Month Template"]

    # Find header row by locating "NAME" in column A, usually row 19 in template
    header_row = None
    for r in range(1, 250):
        if ws.cell(r, 1).value == "NAME":
            header_row = r
            break
    if not header_row:
        raise ValueError("Could not find header row with 'NAME'")

    data_row_start = header_row + 1

    # Write month headers dynamically
    for i, (y, m) in enumerate(months):
       set_cell_value_safe(ws, header_row, MONTH_START_COL + i, month_label(y, m))

    # Decide where STATUS column goes (right after last month)
    status_col = MONTH_START_COL + len(months)  # ignored for now, but kept for layout

    # Determine how far styling should be copied across.
    # Use at least through the last month column (and status col if present in template).
    max_col = max(10, status_col)

    # Snapshot template styles BEFORE inserting rows (insert_rows shifts the template rows down).
    # In your template, row `data_row_start` is the example data row, and `data_row_start + 1` is the TOTAL style row.
    data_row_styles, data_row_height = snapshot_row_style(ws, data_row_start, max_col)
    total_row_styles, total_row_height = snapshot_row_style(ws, data_row_start + 1, max_col)

    # Insert enough rows
    # rows = sum(jobcodes per employee) + total rows per employee
    # compute per employee jobcodes present:
    by_emp = {}
    for key in data.keys():
        by_emp.setdefault(key.employee_id, []).append(key)

    needed = sum(len(keys) + 1 for keys in by_emp.values())
    ws.insert_rows(data_row_start, amount=needed)

    r = data_row_start

    for emp in employees:
        keys = sorted(by_emp.get(emp.id, []), key=lambda k: meta[k]["jobcode"].id)
        block_start = r

        for key in keys:
            emp = meta[key]["employee"]
            jc  = meta[key]["jobcode"]

            # Apply template styling to this newly inserted row
            apply_row_style(ws, r, data_row_styles, data_row_height)

            # A: NAME (choose the best display you have)
            set_cell_value_safe(ws, r, 1, format_employee_name(emp))

            # B: RESOURCE BU
            set_cell_value_safe(ws, r, 2, excel_safe(getattr(emp, "resourceBU", "")))

            # C: CUSTOMER
            set_cell_value_safe(ws, r, 3, excel_safe(getattr(jc, "customerName", "")))

            # D: REPLY ENTITY (FK/object -> readable string)
            reply_entity = getattr(jc, "replyEntity", None)
            set_cell_value_safe(ws, r, 4, excel_safe(reply_entity))

            # E: BUSINESS UNIT
            set_cell_value_safe(ws, r, 5, excel_safe(getattr(jc, "businessUnit", "")))

            # F: JOB ORIGIN
            set_cell_value_safe(ws, r, 6, excel_safe(getattr(jc, "jobOrigin", "")))

            # I: JOB S-ORD CODE
            set_cell_value_safe(ws, r, 9, excel_safe(getattr(jc, "code", "")))

            # J: JOB DESCRIPTION
            set_cell_value_safe(ws, r, 10, excel_safe(getattr(jc, "description", "")))

            # Month cells
            for i, mk in enumerate(months):
                # Keep numeric type so Excel treats it as a number
                set_cell_value_safe(ws, r, MONTH_START_COL + i, data[key].get(mk, Decimal("0")))

            # Status column intentionally not filled for now
            r += 1

        block_end = r - 1

        # TOTAL row
        apply_row_style(ws, r, total_row_styles, total_row_height)

        set_cell_value_safe(ws, r, 2, f"TOTAL - {format_employee_name(emp)}")  # B
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)

        for c in range(MONTH_START_COL, MONTH_START_COL + len(months)):
            col_letter = get_column_letter(c)
            ws.cell(r, c).value = f"=SUM({col_letter}{block_start}:{col_letter}{block_end})"

        # Status column intentionally not filled for now
        r += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="forecast_export.xlsx"'
    return resp