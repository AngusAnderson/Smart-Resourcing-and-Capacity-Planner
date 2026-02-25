from collections import defaultdict, OrderedDict
from dataclasses import dataclass
from datetime import date
from decimal import Decimal

from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from django.db.models import Prefetch
from django.utils.timezone import make_naive

from comwrap.models import ForecastAllocation  # adjust import

# pivot logic to build export data structure

def month_range_keys(start_month: date, end_month: date):
    """Inclusive months. start_month/end_month should be first-of-month dates."""
    y, m = start_month.year, start_month.month
    end_y, end_m = end_month.year, end_month.month
    out = []
    while (y, m) <= (end_y, end_m):
        out.append((y, m))
        m += 1
        if m == 13:
            m = 1
            y += 1
    return out

def add_months(d: date, n: int) -> date:
    y = d.year + (d.month - 1 + n) // 12
    m = (d.month - 1 + n) % 12 + 1
    return date(y, m, 1)

@dataclass(frozen=True)
class ExportRowKey:
    employee_id: int
    jobcode_id: int

def build_forecast_export(start_month: date, end_month: date):
    """
    Returns:
      months: list[(y,m)]
      employees: ordered list of employee objects encountered
      data: dict[(employee_id, jobcode_id)] -> dict[(y,m)] -> Decimal
      meta: dict[(employee_id, jobcode_id)] -> info (employee, jobcode, description, etc.)
    """
    months = month_range_keys(start_month, end_month)
    end_exclusive = add_months(end_month, 1)

    qs = (
        ForecastAllocation.objects
        .filter(
            forecast__date__gte=start_month,
            forecast__date__lt=end_exclusive,
        )
        .select_related(
            "employee",
            "forecast",
            "forecast__jobCode",
        )
        .order_by("employee__id", "forecast__jobCode__id", "forecast__date")
    )

    data = defaultdict(lambda: defaultdict(Decimal))
    meta = {}
    employee_order = OrderedDict()  # preserves first-seen order

    for fa in qs:
        emp = fa.employee
        fc = fa.forecast
        jc = fc.jobCode
        key = ExportRowKey(emp.id, jc.id)
        month_key = (fc.date.year, fc.date.month)

        employee_order.setdefault(emp.id, emp)

        # sum in case there are multiple rows per same key+month
        data[key][month_key] += (fa.daysAllocated or Decimal("0"))

        if key not in meta:
            meta[key] = {
                "employee": emp,
                "jobcode": jc,
                "forecast": fc,  # optional
                # add any fields you want to print in static columns
            }

    return months, list(employee_order.values()), data, meta


# Excel export logic to write the above data structure into the template and return as response

MONTH_START_COL = 11  # column K

def month_label(y, m):
    return date(y, m, 1).strftime("%b-%y")

def export_forecast_xlsx(start_month: date, end_month: date):
    months, employees, data, meta = build_forecast_export(start_month, end_month)

    wb = load_workbook("templates/excel/resource_allocation_template.xlsx")
    ws = wb["Resource Allocation"]

    # Find header row by locating "NAME" in column A (recommended)
    header_row = None
    for r in range(1, 250):
        if ws.cell(r, 1).value == "NAME":
            header_row = r
            break
    if not header_row:
        raise ValueError("Could not find header row with 'NAME'")

    data_row_start = header_row + 1

    # Write month headers dynamically
    for i, (y, m) in enumerate(months):
        ws.cell(header_row, MONTH_START_COL + i).value = month_label(y, m)

    # Decide where STATUS column goes (right after last month)
    status_col = MONTH_START_COL + len(months)

    # Insert enough rows
    # rows = sum(jobcodes per employee) + total rows per employee
    # compute per employee jobcodes present:
    by_emp = {}
    for key in data.keys():
        by_emp.setdefault(key.employee_id, []).append(key)

    needed = sum(len(keys) + 1 for keys in by_emp.values())
    ws.insert_rows(data_row_start, amount=needed)

    style_row = data_row_start + needed  # fallback: or point to a known template style row
    # Better: keep a hidden “sample row” in template at fixed position and copy styles from it.

    r = data_row_start

    for emp in employees:
        keys = sorted(by_emp.get(emp.id, []), key=lambda k: meta[k]["jobcode"].id)
        block_start = r

        for key in keys:
            jc = meta[key]["jobcode"]

            # Static columns (adjust to your template columns)
            ws.cell(r, 1).value = getattr(emp, "name", str(emp))  # A NAME
            ws.cell(r, 9).value = getattr(jc, "code", jc.id)      # I job code
            ws.cell(r, 10).value = getattr(jc, "description", "") # J desc

            # Month cells
            for i, mk in enumerate(months):
                ws.cell(r, MONTH_START_COL + i).value = data[key].get(mk, Decimal("0"))

            ws.cell(r, status_col).value = "O"
            r += 1

        block_end = r - 1

        # TOTAL row
        ws.cell(r, 2).value = f"TOTAL - {getattr(emp, 'name', str(emp))}"  # B
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)

        for c in range(MONTH_START_COL, MONTH_START_COL + len(months)):
            col_letter = get_column_letter(c)
            ws.cell(r, c).value = f"=SUM({col_letter}{block_start}:{col_letter}{block_end})"

        ws.cell(r, status_col).value = "O"
        r += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="forecast_export.xlsx"'
    return resp